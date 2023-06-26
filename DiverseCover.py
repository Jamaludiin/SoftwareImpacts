from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# =========================INSERTING DATA INTO A SHEET OUTOMATICALLY==========================================

wb = load_workbook("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/clean control cc.xlsx")

ws = wb.active

# number of test cases and criteria 
num_test = 8
num_criteria = 4

# variable declarations
num_of_cols = num_criteria + 2  # starts 2 column
num_of_rows = num_test + 1  # number of test cases plus 1 (header row)
num_of_test_case = num_of_rows - 1

# colunm numbers
val_starts_7 = num_of_cols + 1
val_ends_11 = num_of_cols + 5

# CALCULATE N
val_insert_N = num_of_rows + 1
val_range_start_N = 2
val_range_end_N = num_of_rows
for col in range(
    val_range_start_N, num_of_cols
):  
    char = get_column_letter(col)
    # this assigns formula as value to the target cells
    ws[
        char + str(val_insert_N)
    ] = f"=SUM({char + str(val_range_start_N)}:{char + str(val_range_end_N)})"

# give the cell A[num] to a string value
ws["A" + str(val_insert_N)] = "N"


# CALCULATE n(n-1)
val_range_start_n = 2
count = 1
for k in range(num_of_rows - 1):
    count += 1
    j = num_of_cols + 1
    for col in range(val_range_start_n, num_of_cols):
        char = get_column_letter(col)
        char1 = get_column_letter(j)
        ws[char1 + str(count)] = f"={char + str(count)}*({char + str(count)}-1)"
        j += 1

ws[get_column_letter(num_of_cols + 1) + "1"] = "n(n-1)"

ws[get_column_letter(num_of_cols + 1) + "1"].font = Font(bold=True, color="00FF6600")


# CALCULATE Σ n(n-1)
val_insert_sum_n = num_of_rows + 1
val_range_start_n = 2

for col in range(val_starts_7, val_ends_11):
    char = get_column_letter(col)
    ws[
        char + str(val_insert_sum_n)
    ] = f"=SUM({char + str(val_range_start_n)}:{char + str(num_of_rows)})"

ws[get_column_letter(val_starts_7 - 1) + str(val_insert_sum_n)] = "Σ n(n-1)"

# CALCULATE N(N-1)
col2 = 2
for col in range(val_starts_7, val_ends_11):
    char = get_column_letter(col)
    char2 = get_column_letter(col2)
    ws[
        char + str(num_of_rows + 2)
    ] = f"={char2 + str(num_of_rows+1)}*({char2 + str(num_of_rows+1)}-1)"
    col2 += 1

ws[get_column_letter(val_starts_7 - 1) + str(num_of_rows + 2)] = "N(N-1)"


# CALCULATE n(n-1)/N(N-1)
for col in range(val_starts_7, val_ends_11):
    char = get_column_letter(col)
    ws[
        char + str(num_of_rows + 3)
    ] = f"={char + str(num_of_rows+1)}/({char + str(num_of_rows+2)})"

ws[get_column_letter(val_starts_7 - 1) + str(num_of_rows + 3)] = "n(n-1)/N(N-1)"

# CALCULATE 1-D
for col in range(val_starts_7, val_ends_11):
    char = get_column_letter(col)
    ws[char + str(num_of_rows + 4)] = f"=1-{char + str(num_of_rows+3)}"

ws[get_column_letter(val_starts_7 - 1) + str(num_of_rows + 4)] = "1-D"

# True Diversity calculation

j = num_of_cols + num_of_cols
for col in range(1, num_of_rows + 1):
    char = get_column_letter(5)
    char1 = get_column_letter(j)
    char2 = get_column_letter(2)
    ws[char1 + str(col)] = f"=SUM({char2 + str(col)}:{char + str(col)})"

ws[get_column_letter(num_of_cols + num_of_cols) + "1"] = "Sum all n"

ws[get_column_letter(num_of_cols + num_of_cols) + "1"].font = Font(
    bold=True, color="00FF6600"
)


# Σ n(n-1)
j = num_of_cols + num_of_cols + 1
for col in range(1, num_of_rows + 1):
    char = get_column_letter(10)
    char1 = get_column_letter(j)
    char2 = get_column_letter(7)
    ws[char1 + str(col)] = f"=SUM({char2 + str(col)}:{char + str(col)})"

ws[get_column_letter(num_of_cols + num_of_cols + 1) + "1"] = "Σ n(n-1)"

ws[get_column_letter(num_of_cols + num_of_cols + 1) + "1"].font = Font(
    bold=True, color="00FF6600"
)


# N(N-1)
j = num_of_cols + num_of_cols + 2
for col in range(1, num_of_rows + 1):
    # char = get_column_letter(10)
    char1 = get_column_letter(j)
    char2 = get_column_letter(12)
    ws[char1 + str(col)] = f"=({char2 + str(col)}*({char2 + str(col)}-1))"


ws[get_column_letter(num_of_cols + num_of_cols + 2) + "1"] = "N(N-1)"

ws[get_column_letter(num_of_cols + num_of_cols + 2) + "1"].font = Font(
    bold=True, color="00FF6600"
)


# n(n-1)/N(N-1)
j = num_of_cols + num_of_cols + 3
for col in range(1, num_of_rows + 1):
    char = get_column_letter(13)
    char1 = get_column_letter(j)
    char2 = get_column_letter(14)
    if ws[char2 + str(col)].value == "0" and ws[char + str(col)].value == "0":
        ws[char1 + str(col)] = "0"
    else:
        ws[char1 + str(col)] = f"=(({char + str(col)})/({char2 + str(col)}))"

ws[get_column_letter(num_of_cols + num_of_cols + 3) + "1"] = "n(n-1)/N(N-1)"

ws[get_column_letter(num_of_cols + num_of_cols + 3) + "1"].font = Font(
    bold=True, color="00FF6600"
)


# 1-D
j = num_of_cols + num_of_cols + 4
for col in range(1, num_of_rows + 1):
    # char = get_column_letter(13)
    char1 = get_column_letter(j)
    char2 = get_column_letter(15)
    ws[char1 + str(col)] = f"=(1-({char2 + str(col)}))"

ws[get_column_letter(num_of_cols + num_of_cols + 4) + "1"] = "1-D"

ws[get_column_letter(num_of_cols + num_of_cols + 4) + "1"].font = Font(
    bold=True, color="00FF6600"
)

# change headers font styles
F16_I16 = val_starts_7 - 1

for col in range(1, val_starts_7):
    ws[get_column_letter(col) + "1"].font = Font(bold=True, color="00FF6600")
    ws[get_column_letter(col) + str(num_of_rows + 1)].font = Font(
        bold=True, color="00FF6600"
    )
    # ws[get_column_letter(col) + '16'].font = Font(bold=True,color='00FF6600')
    ws[get_column_letter(F16_I16 - 1) + str(num_of_rows + 1)].font = Font(
        bold=True, color="00FF6600"
    )
    ws[get_column_letter(F16_I16 - 1) + str(num_of_rows + 2)].font = Font(
        bold=True, color="00FF6600"
    )
    ws[get_column_letter(F16_I16 - 1) + str(num_of_rows + 3)].font = Font(
        bold=True, color="00FF6600"
    )
    ws[get_column_letter(F16_I16 - 1) + str(num_of_rows + 4)].font = Font(
        bold=True, color="00FF6600"
    )
    ws[get_column_letter(F16_I16) + str(num_of_rows + 5)].font = Font(
        bold=True, color="00FF6600"
    )

    F16_I16 += 1
wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/clean control cc.xlsx")
